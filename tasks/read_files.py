import os
from datetime import datetime, timedelta
from uuid import UUID

from logger import logger
from openpyxl import load_workbook

relative_path = '../admin/Menu.xlsx'
current_directory = os.path.dirname(__file__)
PATH_FILE_EXCEL = os.path.join(current_directory, relative_path)


def is_change_file(file_path: str,
                   threshold_seconds: int = 15):

    file_stat = os.stat(file_path)
    last_modification_time = datetime.fromtimestamp(file_stat.st_mtime)
    delta_time = datetime.now() - last_modification_time
    return delta_time <= timedelta(seconds=threshold_seconds)


def is_uuid(data_str):
    try:
        UUID(data_str)
        return True
    except ValueError:
        logger.info('Not UUID')
    return False


def get_data_from_excel_file(excel_file_path=PATH_FILE_EXCEL):
    try:
        logger.info('Start reading file')
        workbook = load_workbook(excel_file_path, data_only=True)
        worksheet = workbook.active
        menus = dict()
        submenus = dict()
        dishes = dict()
        current_menu_id = None
        current_submenu_id = None
        for row in worksheet.iter_rows(values_only=True):
            # logger.info(f"Current row {row}")
            if row[0] and is_uuid(row[0]):
                menus[UUID(row[0])] = {
                    'title': row[1],
                    'description': row[2]
                }
                current_menu_id = UUID(row[0])
                current_submenu_id = None
            if row[1] and is_uuid(row[1]):
                submenus[UUID(row[1])] = {
                    'title': row[2],
                    'description': row[3],
                    'menu_id': current_menu_id
                }
                current_submenu_id = UUID(row[1])
            if row[2] and is_uuid(row[2]):
                if row[6]:
                    discount = int(row[6])
                else:
                    discount = 0
                dishes[UUID(row[2])] = {
                    'title': row[3],
                    'description': row[4],
                    'price': str(row[5]),
                    'discount': discount,
                    'submenu_id': current_submenu_id,
                    'menu_id': current_menu_id
                }
        # logger.info(f'Row data menu {menus}')
        # logger.info(f'Row data submenu {submenus}')
        # logger.info(f'Row data dish {dishes}')
        logger.info('File is reading')
        return menus, submenus, dishes
    except Exception as ex:
        logger.error(f'ERRROR!!!!!!! {ex}')
    finally:
        workbook.close()
